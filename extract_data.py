import pandas as pd
import openpyxl
import json

file_path = 'Grille 2025 (1).xlsx'

print("=" * 80)
print("EXTRACTION DES DONNÉES STRUCTURÉES")
print("=" * 80)

# Charger le classeur
wb = openpyxl.load_workbook(file_path, data_only=True)

# 1. ANALYSER LA FEUILLE "grille_de_cotation"
print("\n" + "="*80)
print("1. STRUCTURE DE LA GRILLE DE COTATION")
print("="*80)

ws_grille = wb['grille_de_cotation']
df_grille = pd.read_excel(file_path, sheet_name='grille_de_cotation', header=None)

# Trouver les catégories principales
categories = []
current_category = None
items = []

for idx, row in df_grille.iterrows():
    col0 = str(row[0]) if pd.notna(row[0]) else ""
    
    # Détecter une catégorie principale (commence par un chiffre suivi d'un point)
    if any(col0.startswith(f"{i}.") for i in range(1, 10)):
        if current_category:
            categories.append(current_category)
        current_category = {
            'name': col0,
            'items': []
        }
    # Détecter un sous-item (ligne avec KO = 1 dans colonne 2)
    elif current_category and pd.notna(row[2]) and str(row[2]) == '1':
        item_name = col0
        ponderation = row[8] if pd.notna(row[8]) else None
        note = row[9] if pd.notna(row[9]) else None
        
        # Chercher les commentaires dans les lignes suivantes
        comments = []
        actions = []
        for next_idx in range(idx + 1, min(idx + 10, len(df_grille))):
            next_row = df_grille.iloc[next_idx]
            if pd.notna(next_row[10]):  # Colonne commentaires
                comments.append(str(next_row[10]))
            if pd.notna(next_row[11]):  # Colonne actions correctives
                actions.append(str(next_row[11]))
            # Arrêter si on trouve une nouvelle catégorie ou item
            if pd.notna(next_row[0]) and (any(str(next_row[0]).startswith(f"{i}.") for i in range(1, 10)) or pd.notna(next_row[2])):
                break
        
        current_category['items'].append({
            'name': item_name,
            'ponderation': float(ponderation) if ponderation else None,
            'note': float(note) if note else None,
            'comments': comments,
            'actions': actions
        })

if current_category:
    categories.append(current_category)

print(f"\nNombre de catégories trouvées: {len(categories)}")
for cat in categories:
    print(f"\n  {cat['name']}")
    print(f"    Nombre d'items: {len(cat['items'])}")
    for item in cat['items'][:2]:  # Afficher les 2 premiers items
        print(f"      - {item['name']}")
        print(f"        Pondération: {item['ponderation']}, Note: {item['note']}")

# 2. ANALYSER LA FEUILLE "Listes_défilantes"
print("\n" + "="*80)
print("2. STRUCTURE DES LISTES DÉFILANTES")
print("="*80)

df_listes = pd.read_excel(file_path, sheet_name='Listes_défilantes', header=None)

# Structure: Colonne 0 = Catégorie/Item, Colonne 1 = Observation, Colonne 2 = Action
observations = {}
current_item = None

for idx, row in df_listes.iterrows():
    col0 = str(row[0]) if pd.notna(row[0]) else ""
    col1 = str(row[1]) if pd.notna(row[1]) else ""
    col2 = str(row[2]) if pd.notna(row[2]) else ""
    
    # Nouvelle catégorie ou item
    if col0 and col0 != 'nan':
        current_item = col0
        if current_item not in observations:
            observations[current_item] = []
    
    # Observation avec action
    if current_item and col1 and col1 != 'nan':
        observations[current_item].append({
            'observation': col1,
            'action': col2 if col2 != 'nan' else None
        })

print(f"\nNombre d'items avec observations: {len(observations)}")
for item, obs_list in list(observations.items())[:3]:
    print(f"\n  {item}")
    print(f"    Nombre d'observations: {len(obs_list)}")
    for obs in obs_list[:2]:
        print(f"      - {obs['observation']}")

# 3. ANALYSER LA FEUILLE "cartographie"
print("\n" + "="*80)
print("3. STRUCTURE DE LA CARTOGRAPHIE")
print("="*80)

ws_carto = wb['cartographie']
print(f"Dimensions: {ws_carto.max_row} lignes x {ws_carto.max_column} colonnes")

# Chercher les formules de calcul
print("\nRecherche des formules de calcul...")
for row in range(1, min(50, ws_carto.max_row + 1)):
    for col in range(1, min(15, ws_carto.max_column + 1)):
        cell = ws_carto.cell(row, col)
        if cell.data_type == 'f':  # Formule
            print(f"  Ligne {row}, Colonne {col}: {cell.formula}")

# 4. EXPORTER EN JSON pour faciliter le développement
print("\n" + "="*80)
print("4. EXPORT DES DONNÉES STRUCTURÉES")
print("="*80)

data_structure = {
    'categories': categories,
    'observations': observations
}

with open('data_structure.json', 'w', encoding='utf-8') as f:
    json.dump(data_structure, f, ensure_ascii=False, indent=2)

print("\n[OK] Donnees exportees dans 'data_structure.json'")
print(f"   - {len(categories)} catégories")
print(f"   - {sum(len(cat['items']) for cat in categories)} items au total")
print(f"   - {len(observations)} items avec observations")

print("\n" + "="*80)
print("EXTRACTION TERMINÉE")
print("="*80)

