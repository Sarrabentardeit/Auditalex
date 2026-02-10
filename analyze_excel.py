import openpyxl
import json

# Charger le fichier Excel
wb = openpyxl.load_workbook('Grille 2025 (1).xlsx', data_only=True)

# Trouver l'onglet grille_de_cotation
sheet_name = None
for name in wb.sheetnames:
    if 'grille' in name.lower() or 'cotation' in name.lower():
        sheet_name = name
        break

if not sheet_name:
    sheet_name = wb.sheetnames[0]

print(f"Analyse de l'onglet: {sheet_name}")
ws = wb[sheet_name]

# Analyser la structure
print(f"\nDimensions: {ws.max_row} lignes x {ws.max_column} colonnes")

# Analyser les premières lignes pour comprendre la structure
print("\n=== Structure des colonnes ===")
header_row = None
for row_idx, row in enumerate(ws.iter_rows(max_row=50), 1):
    row_values = [cell.value for cell in row[:15]]
    if any(v and str(v).strip() for v in row_values):
        print(f"Ligne {row_idx}: {row_values}")
        if row_idx == 1 or (header_row is None and any('catégorie' in str(v).lower() or 'item' in str(v).lower() or 'note' in str(v).lower() for v in row_values if v)):
            header_row = row_idx

# Analyser les catégories et items
print("\n=== Catégories et Items ===")
categories = []
current_category = None
current_items = []

for row_idx, row in enumerate(ws.iter_rows(min_row=header_row or 1, max_row=ws.max_row), header_row or 1):
    row_values = [cell.value for cell in row[:15]]
    
    # Chercher les catégories (généralement en gras ou avec un numéro)
    cat_cell = row[0] if len(row) > 0 else None
    item_cell = row[1] if len(row) > 1 else None
    
    if cat_cell and cat_cell.value:
        cat_value = str(cat_cell.value).strip()
        # Détecter une catégorie (commence par un numéro ou contient "LOCAUX", "MAITRISE", etc.)
        if (cat_value[0].isdigit() and '.' in cat_value) or any(keyword in cat_value.upper() for keyword in ['LOCAUX', 'MAITRISE', 'TRACABILITE', 'GESTION', 'DECHETS', 'PERSONNEL']):
            if current_category:
                categories.append({
                    'name': current_category,
                    'items': current_items
                })
            current_category = cat_value
            current_items = []
            print(f"\nCatégorie trouvée: {cat_value}")
    
    # Chercher les items (généralement dans la colonne suivante)
    if item_cell and item_cell.value and current_category:
        item_value = str(item_cell.value).strip()
        # Détecter un item (pas vide, pas un numéro seul)
        if item_value and not item_value.isdigit() and len(item_value) > 3:
            # Chercher la pondération (généralement dans une colonne spécifique)
            ponderation = None
            note = None
            classification = None
            
            # Analyser les colonnes pour trouver pondération, note, etc.
            for col_idx, cell in enumerate(row[2:10], 2):
                if cell.value is not None:
                    val = cell.value
                    if isinstance(val, (int, float)):
                        if 0 < val < 1 and ponderation is None:
                            ponderation = val
                        elif val in [0, 0.3, 0.7, 1.0] and note is None:
                            note = val
            
            item_data = {
                'name': item_value,
                'ponderation': ponderation,
                'note': note,
                'row': row_idx
            }
            current_items.append(item_data)
            print(f"  - Item: {item_value[:50]}... (pond: {ponderation}, note: {note})")

# Ajouter la dernière catégorie
if current_category:
    categories.append({
        'name': current_category,
        'items': current_items
    })

print(f"\n=== Résumé ===")
print(f"Nombre de catégories trouvées: {len(categories)}")
for cat in categories:
    print(f"  {cat['name']}: {len(cat['items'])} items")

# Sauvegarder pour comparaison
with open('excel_structure.json', 'w', encoding='utf-8') as f:
    json.dump(categories, f, ensure_ascii=False, indent=2)

print("\nStructure sauvegardée dans excel_structure.json")
